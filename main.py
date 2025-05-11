import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sqlite3
from datetime import datetime
from tkcalendar import DateEntry
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import os

class VegetableInventory:
    def __init__(self, root):
        self.root = root
        self.root.title("蔬菜批发进销存管理系统")
        self.root.geometry("1000x700")  # 调整窗口大小
        self.root.resizable(False, False)  # 禁止调整窗口大小
        
        # 设置主题颜色
        self.style = ttk.Style()
        self.style.configure("TFrame", background="#f0f0f0")
        self.style.configure("TLabelframe", background="#f0f0f0")
        self.style.configure("TLabelframe.Label", font=("微软雅黑", 10, "bold"))
        self.style.configure("TButton", font=("微软雅黑", 9))
        self.style.configure("TLabel", font=("微软雅黑", 9))
        
        # 创建数据库连接
        self.conn = sqlite3.connect('vegetable_inventory.db')
        self.cursor = self.conn.cursor()
        
        # 创建数据库表
        self.create_tables()
        
        # 创建主框架
        self.main_frame = ttk.Frame(self.root, padding="5", style="TFrame")
        self.main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 创建顶部统计面板
        self.create_stats_panel()
        
        # 创建输入框架
        self.create_input_frame()
        
        # 创建表格
        self.create_data_tables()
        
        # 加载数据
        self.load_data()
        
    def create_tables(self):
        # 创建进货表
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS purchases (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                product_name TEXT NOT NULL,
                purchase_date TEXT NOT NULL,
                purchase_price REAL NOT NULL,
                quantity REAL NOT NULL,
                total_amount REAL NOT NULL
            )
        ''')
        
        # 创建销售表
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS sales (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                product_name TEXT NOT NULL,
                sale_date TEXT NOT NULL,
                sale_price REAL NOT NULL,
                quantity REAL NOT NULL,
                total_amount REAL NOT NULL,
                profit REAL NOT NULL
            )
        ''')
        
        self.conn.commit()
        
    def create_stats_panel(self):
        stats_frame = ttk.LabelFrame(self.main_frame, text="库存概览", padding="5")
        stats_frame.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=2)
        
        # 创建统计信息标签
        self.total_products_label = ttk.Label(stats_frame, text="产品总数: 0")
        self.total_products_label.grid(row=0, column=0, padx=10)
        
        self.total_purchase_label = ttk.Label(stats_frame, text="总进货金额: ¥0.00")
        self.total_purchase_label.grid(row=0, column=1, padx=10)
        
        self.total_sales_label = ttk.Label(stats_frame, text="总销售金额: ¥0.00")
        self.total_sales_label.grid(row=0, column=2, padx=10)
        
        self.total_profit_label = ttk.Label(stats_frame, text="总利润: ¥0.00")
        self.total_profit_label.grid(row=0, column=3, padx=10)
        
        self.low_stock_label = ttk.Label(stats_frame, text="库存预警: 0个产品")
        self.low_stock_label.grid(row=0, column=4, padx=10)
        
    def create_input_frame(self):
        self.input_frame = ttk.LabelFrame(self.main_frame, text="数据输入", padding="5")
        self.input_frame.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=2)
        
        # 产品名称（下拉选择框）
        ttk.Label(self.input_frame, text="产品名称:").grid(row=0, column=0, sticky=tk.W, padx=2)
        self.product_names = [
            "空心菜", "水白菜", "水萝卜", "油麦菜", "菜心",
            "塔菜", "白萝卜", "快白菜", "小白菜", "大白菜"
        ]
        self.product_name = ttk.Combobox(self.input_frame, values=self.product_names, state="readonly", width=13, font=("微软雅黑", 9))
        self.product_name.grid(row=0, column=1, sticky=tk.W, padx=2)
        
        # 日期选择
        ttk.Label(self.input_frame, text="日期:").grid(row=0, column=2, sticky=tk.W, padx=2)
        self.date_entry = DateEntry(self.input_frame, width=10, background='darkblue',
                                  foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd',
                                  font=("微软雅黑", 9))
        self.date_entry.grid(row=0, column=3, sticky=tk.W, padx=2)
        
        # 价格和数量
        ttk.Label(self.input_frame, text="单价:").grid(row=1, column=0, sticky=tk.W, padx=2)
        self.price = ttk.Entry(self.input_frame, width=15, font=("微软雅黑", 9))
        self.price.grid(row=1, column=1, sticky=tk.W, padx=2)
        
        ttk.Label(self.input_frame, text="数量:").grid(row=1, column=2, sticky=tk.W, padx=2)
        self.quantity = ttk.Entry(self.input_frame, width=15, font=("微软雅黑", 9))
        self.quantity.grid(row=1, column=3, sticky=tk.W, padx=2)
        
        # 操作按钮
        self.button_frame = ttk.Frame(self.input_frame)
        self.button_frame.grid(row=2, column=0, columnspan=4, pady=5)
        
        ttk.Button(self.button_frame, text="添加进货", command=self.add_purchase,
                  style="TButton", width=10).pack(side=tk.LEFT, padx=2)
        ttk.Button(self.button_frame, text="添加销售", command=self.add_sale,
                  style="TButton", width=10).pack(side=tk.LEFT, padx=2)
        ttk.Button(self.button_frame, text="清空", command=self.clear_inputs,
                  style="TButton", width=10).pack(side=tk.LEFT, padx=2)
        ttk.Button(self.button_frame, text="导出Excel", command=self.export_to_excel,
                  style="TButton", width=10).pack(side=tk.LEFT, padx=2)
        
    def create_data_tables(self):
        # 创建进货记录表格
        self.purchase_frame = ttk.LabelFrame(self.main_frame, text="进货记录", padding="5")
        self.purchase_frame.grid(row=2, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), pady=2)
        
        # 设置表格样式
        style = ttk.Style()
        style.configure("Treeview", font=("微软雅黑", 9))
        style.configure("Treeview.Heading", font=("微软雅黑", 9, "bold"))
        
        self.purchase_tree = ttk.Treeview(self.purchase_frame, columns=("id", "product", "date", "price", "quantity", "total"),
                                        show="headings", height=15)
        self.purchase_tree.heading("id", text="ID")
        self.purchase_tree.heading("product", text="产品名称")
        self.purchase_tree.heading("date", text="进货日期")
        self.purchase_tree.heading("price", text="进货价")
        self.purchase_tree.heading("quantity", text="数量")
        self.purchase_tree.heading("total", text="总金额")
        
        self.purchase_tree.column("id", width=40, anchor="center")
        self.purchase_tree.column("product", width=120, anchor="center")
        self.purchase_tree.column("date", width=80, anchor="center")
        self.purchase_tree.column("price", width=80, anchor="center")
        self.purchase_tree.column("quantity", width=80, anchor="center")
        self.purchase_tree.column("total", width=80, anchor="center")
        
        self.purchase_tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 创建销售记录表格
        self.sale_frame = ttk.LabelFrame(self.main_frame, text="销售记录", padding="5")
        self.sale_frame.grid(row=2, column=1, sticky=(tk.W, tk.E, tk.N, tk.S), pady=2)
        
        self.sale_tree = ttk.Treeview(self.sale_frame, columns=("id", "product", "date", "price", "quantity", "total", "profit"),
                                    show="headings", height=15)
        self.sale_tree.heading("id", text="ID")
        self.sale_tree.heading("product", text="产品名称")
        self.sale_tree.heading("date", text="销售日期")
        self.sale_tree.heading("price", text="销售价")
        self.sale_tree.heading("quantity", text="数量")
        self.sale_tree.heading("total", text="总金额")
        self.sale_tree.heading("profit", text="利润")
        
        self.sale_tree.column("id", width=40, anchor="center")
        self.sale_tree.column("product", width=120, anchor="center")
        self.sale_tree.column("date", width=80, anchor="center")
        self.sale_tree.column("price", width=80, anchor="center")
        self.sale_tree.column("quantity", width=80, anchor="center")
        self.sale_tree.column("total", width=80, anchor="center")
        self.sale_tree.column("profit", width=80, anchor="center")
        
        self.sale_tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 添加滚动条
        purchase_scroll = ttk.Scrollbar(self.purchase_frame, orient=tk.VERTICAL, command=self.purchase_tree.yview)
        purchase_scroll.grid(row=0, column=1, sticky=(tk.N, tk.S))
        self.purchase_tree.configure(yscrollcommand=purchase_scroll.set)
        
        sale_scroll = ttk.Scrollbar(self.sale_frame, orient=tk.VERTICAL, command=self.sale_tree.yview)
        sale_scroll.grid(row=0, column=1, sticky=(tk.N, tk.S))
        self.sale_tree.configure(yscrollcommand=sale_scroll.set)
        
        # 配置grid权重，使表格可以随窗口调整大小
        self.main_frame.grid_columnconfigure(0, weight=1)
        self.main_frame.grid_columnconfigure(1, weight=1)
        self.main_frame.grid_rowconfigure(2, weight=1)
        
        self.purchase_frame.grid_columnconfigure(0, weight=1)
        self.purchase_frame.grid_rowconfigure(0, weight=1)
        
        self.sale_frame.grid_columnconfigure(0, weight=1)
        self.sale_frame.grid_rowconfigure(0, weight=1)
        
    def update_stats(self):
        # 更新产品总数
        self.cursor.execute('SELECT COUNT(DISTINCT product_name) FROM purchases')
        total_products = self.cursor.fetchone()[0]
        self.total_products_label.config(text=f"产品总数: {total_products}")
        
        # 更新总进货金额
        self.cursor.execute('SELECT SUM(total_amount) FROM purchases')
        total_purchase = self.cursor.fetchone()[0] or 0
        self.total_purchase_label.config(text=f"总进货金额: ¥{total_purchase:.2f}")
        
        # 更新总销售金额和利润
        self.cursor.execute('SELECT SUM(total_amount), SUM(profit) FROM sales')
        total_sales, total_profit = self.cursor.fetchone()
        total_sales = total_sales or 0
        total_profit = total_profit or 0
        self.total_sales_label.config(text=f"总销售金额: ¥{total_sales:.2f}")
        self.total_profit_label.config(text=f"总利润: ¥{total_profit:.2f}")
        
        # 更新库存预警
        self.cursor.execute('''
            SELECT COUNT(*) FROM (
                SELECT product_name, 
                       SUM(CASE WHEN type='purchase' THEN quantity ELSE -quantity END) as stock
                FROM (
                    SELECT product_name, quantity, 'purchase' as type FROM purchases
                    UNION ALL
                    SELECT product_name, quantity, 'sale' as type FROM sales
                )
                GROUP BY product_name
                HAVING stock < 10
            )
        ''')
        low_stock_count = self.cursor.fetchone()[0]
        self.low_stock_label.config(text=f"库存预警: {low_stock_count}个产品")
        
    def add_purchase(self):
        try:
            product = self.product_name.get().strip()
            date = self.date_entry.get()
            price = float(self.price.get())
            quantity = float(self.quantity.get())
            
            if not product:
                messagebox.showerror("错误", "请输入产品名称")
                return
                
            total = price * quantity
            
            self.cursor.execute('''
                INSERT INTO purchases (product_name, purchase_date, purchase_price, quantity, total_amount)
                VALUES (?, ?, ?, ?, ?)
            ''', (product, date, price, quantity, total))
            
            self.conn.commit()
            self.load_data()
            self.clear_inputs()
            messagebox.showinfo("成功", "进货记录已添加")
            
        except ValueError:
            messagebox.showerror("错误", "请输入有效的数字")
        except Exception as e:
            messagebox.showerror("错误", str(e))
            
    def add_sale(self):
        try:
            product = self.product_name.get().strip()
            date = self.date_entry.get()
            price = float(self.price.get())
            quantity = float(self.quantity.get())
            
            if not product:
                messagebox.showerror("错误", "请输入产品名称")
                return
                
            # 检查库存
            self.cursor.execute('''
                SELECT SUM(quantity) FROM purchases 
                WHERE product_name = ?
            ''', (product,))
            total_purchased = self.cursor.fetchone()[0] or 0
            
            self.cursor.execute('''
                SELECT SUM(quantity) FROM sales 
                WHERE product_name = ?
            ''', (product,))
            total_sold = self.cursor.fetchone()[0] or 0
            
            available_stock = total_purchased - total_sold
            
            if quantity > available_stock:
                messagebox.showerror("错误", f"库存不足！当前库存: {available_stock}")
                return
                
            # 获取最近一次进货价格
            self.cursor.execute('''
                SELECT purchase_price FROM purchases 
                WHERE product_name = ? 
                ORDER BY purchase_date DESC LIMIT 1
            ''', (product,))
            purchase_price = self.cursor.fetchone()
            
            if not purchase_price:
                messagebox.showerror("错误", "未找到该产品的进货记录")
                return
                
            purchase_price = purchase_price[0]
            total = price * quantity
            profit = total - (purchase_price * quantity)
            
            self.cursor.execute('''
                INSERT INTO sales (product_name, sale_date, sale_price, quantity, total_amount, profit)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (product, date, price, quantity, total, profit))
            
            self.conn.commit()
            self.load_data()
            self.clear_inputs()
            messagebox.showinfo("成功", "销售记录已添加")
            
        except ValueError:
            messagebox.showerror("错误", "请输入有效的数字")
        except Exception as e:
            messagebox.showerror("错误", str(e))
            
    def load_data(self):
        # 清空现有数据
        for item in self.purchase_tree.get_children():
            self.purchase_tree.delete(item)
        for item in self.sale_tree.get_children():
            self.sale_tree.delete(item)
            
        # 加载进货记录
        self.cursor.execute('SELECT * FROM purchases ORDER BY purchase_date DESC')
        for row in self.cursor.fetchall():
            self.purchase_tree.insert('', 'end', values=row)
            
        # 加载销售记录
        self.cursor.execute('SELECT * FROM sales ORDER BY sale_date DESC')
        for row in self.cursor.fetchall():
            self.sale_tree.insert('', 'end', values=row)
            
        # 更新统计信息
        self.update_stats()
            
    def clear_inputs(self):
        self.product_name.set("")
        self.price.delete(0, tk.END)
        self.quantity.delete(0, tk.END)
        self.date_entry.set_date(datetime.now())
        
    def export_to_excel(self):
        try:
            # 创建新的Excel工作簿
            wb = openpyxl.Workbook()
            
            # 创建进货记录表
            ws_purchase = wb.active
            ws_purchase.title = "进货记录"
            
            # 设置表头
            headers = ["ID", "产品名称", "进货日期", "进货价", "数量", "总金额"]
            for col, header in enumerate(headers, 1):
                cell = ws_purchase.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # 写入进货数据
            self.cursor.execute('SELECT * FROM purchases ORDER BY purchase_date DESC')
            for row_idx, row in enumerate(self.cursor.fetchall(), 2):
                for col_idx, value in enumerate(row, 1):
                    cell = ws_purchase.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.alignment = Alignment(horizontal="center")
            
            # 创建销售记录表
            ws_sale = wb.create_sheet(title="销售记录")
            
            # 设置表头
            headers = ["ID", "产品名称", "销售日期", "销售价", "数量", "总金额", "利润"]
            for col, header in enumerate(headers, 1):
                cell = ws_sale.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # 写入销售数据
            self.cursor.execute('SELECT * FROM sales ORDER BY sale_date DESC')
            for row_idx, row in enumerate(self.cursor.fetchall(), 2):
                for col_idx, value in enumerate(row, 1):
                    cell = ws_sale.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.alignment = Alignment(horizontal="center")
            
            # 调整列宽
            for ws in [ws_purchase, ws_sale]:
                for column in ws.columns:
                    max_length = 0
                    column = [cell for cell in column]
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column[0].column_letter].width = adjusted_width
            
            # 保存文件
            filename = f"蔬菜批发进销存记录_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            wb.save(filename)
            messagebox.showinfo("成功", f"数据已导出到 {filename}")
            
        except Exception as e:
            messagebox.showerror("错误", f"导出失败: {str(e)}")

if __name__ == "__main__":
    root = tk.Tk()
    app = VegetableInventory(root)
    root.mainloop()